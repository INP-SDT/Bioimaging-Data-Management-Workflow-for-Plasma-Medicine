# -*- coding: utf-8 -*-
"""
Created on Wed May 15 14:13:49 2024

@author: WagnerR
"""
# imports for ELN API handling
import elabapi_python
from elabapi_python.rest import ApiException
import warnings

# imports for user inputs
import tkinter as tk
from tkinter import simpledialog

# imports for data import/export to user infrastructure
import sys
import json
from typing import Literal
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook

# %% Individual configuration of the API accoring the individual api key given as a string by the user
"""
To use this handler script please exhange 'hostAdress' by the url of your eLabFTW instance
"""
def API_configurator_direct_input(apiKey):
    # suppress occuring warnings during api configuration
    warnings.filterwarnings("ignore")  
    
    #########################
    #         CONFIG        #
    #########################
    configuration = elabapi_python.Configuration()
    # catch exception if no api key file was provided
    try:
        configuration.api_key['api_key'] = apiKey
    except UnboundLocalError:
        print("No txt file found in the directory.")
        sys.exit()
    configuration.api_key_prefix['api_key'] = 'Authorization'
    configuration.host = 'hostAdress'
    configuration.debug = False
    configuration.verify_ssl = False
    #########################
    #      END CONFIG       #
    #########################

    # create an instance of the API class
    api_client = elabapi_python.ApiClient(configuration)
    # fix issue with Authorization header not being properly set by the generated lib
    api_client.set_default_header(header_name='Authorization', header_value=apiKey)
    
    # build api instances by calling the build_api_instances function
    build_api_instances(api_client)
    API_test()    

# %% Creation of API instances for usage based on a client, which is returned by API_configurator()
def build_api_instances(client):
    # API instance of the API keys 
    global KeyApi
    KeyApi = elabapi_python.ApiKeysApi(client)
    
    # API instance for the access to the available experiments 
    global experimentApi
    experimentApi = elabapi_python.ExperimentsApi(client)
    #print("Experiment API is build")
    
    global uploadApi
    uploadApi = elabapi_python.UploadsApi(client)
    #print("Attachment API is build")

# %% Test for API key validity
def API_test():
    try: 
        # Read API keys
        KeyApi.get_apikeys()
        print("API key is valid.\nAccess to eLabFTW granted")
    except ApiException:
        print("Given API key is not valid!")
    
# %% extract attachments for a given experiment
def extract_experiment_attachments(expID: int = None, experiment = None, export = False, filetype: Literal["JSON metadata", "Excel file"] = "JSON metadata"):
    # 1. case: ELN experiment is given but not the expID 
    
    if filetype not in ["JSON metadata", "Excel file"]:
        raise ValueError("Invalid type. Must be 'JSON metadata' or 'Excel file'.")
    if expID is None and experiment is not None:
        attachments = uploadApi.read_uploads('experiments' , experiment.id)
    else:
        # 2. case: not expID is given but no experiment is given
        if expID is None and experiment is None:
            root = tk.Tk()
            root.withdraw()
            expID = simpledialog.askstring(title="Input", prompt="Enter the experiment ID of your experiment:")
            attachments = uploadApi.read_uploads('experiments' , expID)
        else:
            attachments = uploadApi.read_uploads('experiments' , expID)
    # loop through the attached files and check if the given file with the given title is already existing return the link for the ELN entry
    for attachment in attachments:
        try:
            if filetype == "JSON metadata":
                if attachment.real_name == "json_data.json":
                    item = uploadApi.read_upload("experiments", expID, attachment.id, format='binary', _preload_content=False)
                    # store the utf-8 decoded output of the extracted 
                    output = item.data.decode("utf-8")
                    json_data = json.loads(output)
                    if export == True:
                        with open('output.json', 'w') as file:
                            # Write the dictionary object to the file
                            json.dump(json_data, file, indent=4)
                    else:
                        return json_data
            else:
                if attachment.real_name == "biologicalMetadata.xlsx":
                    # extract uploaded Excel file as binary
                    item = uploadApi.read_upload("experiments", expID, attachment.id, format='binary', _preload_content=False) 
                    # put binary file into bytes
                    excelData = BytesIO(item.data)
                    frameWithExcelData = load_excel_data(excelData)
                    return frameWithExcelData
        except: 
            continue
# %% load attached excel files into a dataframe
def load_excel_data(file):
    try:
        # Load the workbook
        workbook = load_workbook(file)

        # Select the active sheet (or specify a sheet name)
        sheet = workbook.active  # Or workbook["SheetName"]

        # Read data from the sheet
        data = []
        for row in sheet.iter_rows(values_only=True):
            data.append(row)

    except FileNotFoundError:
        print(f"Error: The file at {file} was not found.")
    except Exception as e:
        print(f"An error occurred: {e}")

    # Create a DataFrame
    df = pd.DataFrame(data)
    return df